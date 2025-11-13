from openpyxl import load_workbook
from openpyxl.styles.fills import PatternFill

# Load workbook
wb = load_workbook('Mobile_App_Security_Checklist_es.xlsx')
ws = wb['Security Requirements']

print('=== ANALISIS COMPLETO DEL CHECKLIST OWASP MASVS ===\n')

# Find all MSTG controls
controls = []
current_category = ""

for i, row in enumerate(ws.iter_rows(min_row=1, max_row=200), 1):
    # Check for category in column B
    if row[1].value and 'Requirements' in str(row[1].value):
        current_category = row[1].value
        print(f'\n{"="*80}')
        print(f'CATEGORÍA: {current_category}')
        print(f'{"="*80}\n')
        continue

    # Get MSTG ID from column C (index 2)
    mstg_id = row[2].value if len(row) > 2 else None
    desc = row[3].value if len(row) > 3 else None  # Description in column D
    status = row[9].value if len(row) > 9 else None  # Column J (Status)

    if mstg_id and str(mstg_id).startswith('MSTG-'):
        # Check all cells in the row for green background
        is_green = False
        colors_found = []

        for idx, cell in enumerate(row[:10]):
            if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                try:
                    color = str(cell.fill.start_color.rgb) if hasattr(cell.fill.start_color, 'rgb') else str(cell.fill.start_color.index)
                    colors_found.append(f'Col{idx}:{color}')
                    # Common green colors in Excel
                    if any(green in color.upper() for green in ['00B050', '92D050', 'C6E0B4', '70AD47', '00FF00']):
                        is_green = True
                except:
                    pass

        control_info = {
            'id': mstg_id,
            'desc': desc,
            'category': current_category,
            'status': status,
            'is_green': is_green,
            'row': i,
            'colors': colors_found
        }
        controls.append(control_info)

        marker = '✓ VERDE (PRIORIDAD)' if is_green else ''
        print(f'{mstg_id} {marker}')
        if desc:
            desc_short = str(desc)[:100] + '...' if len(str(desc)) > 100 else desc
            print(f'   {desc_short}')
        if is_green and colors_found:
            print(f'   Colors: {", ".join(colors_found[:3])}')
        print()

print(f'\n{"="*80}')
print(f'RESUMEN')
print(f'{"="*80}')
print(f'Total controles encontrados: {len(controls)}')
green_controls = [c for c in controls if c['is_green']]
print(f'Controles VERDES (Prioridad): {len(green_controls)}')

if green_controls:
    print(f'\n{"="*80}')
    print(f'LISTA DE CONTROLES VERDES (IMPLEMENTACIÓN PRIORITARIA)')
    print(f'{"="*80}\n')
    for ctrl in green_controls:
        desc_text = ctrl["desc"][:70] + '...' if ctrl["desc"] and len(str(ctrl["desc"])) > 70 else ctrl["desc"]
        print(f'{ctrl["id"]}: {desc_text}')

# If no green controls found, show sample of colors to debug
if len(green_controls) == 0:
    print(f'\n{"="*80}')
    print('DEBUG: Mostrando colores encontrados en las primeras 10 filas MSTG')
    print(f'{"="*80}\n')
    for ctrl in controls[:10]:
        if ctrl['colors']:
            print(f'{ctrl["id"]}: {ctrl["colors"]}')
