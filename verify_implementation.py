from openpyxl import load_workbook
import re

# Load workbook to get green controls
wb = load_workbook('Mobile_App_Security_Checklist_es.xlsx')
ws = wb['Security Requirements']

# Get green controls
green_controls = []
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=200), 1):
    mstg_id = row[2].value if len(row) > 2 else None
    desc = row[3].value if len(row) > 3 else None

    if mstg_id and str(mstg_id).startswith('MSTG-'):
        if len(row) > 5 and row[5].fill and hasattr(row[5].fill, 'start_color') and row[5].fill.start_color:
            try:
                color = str(row[5].fill.start_color.rgb) if hasattr(row[5].fill.start_color, 'rgb') else str(row[5].fill.start_color.index)
                if '0099CC00' in color or '99CC00' in color:
                    green_controls.append(mstg_id)
            except:
                pass

# Read documento_owasp.tex
with open('documento_owasp.tex', 'r', encoding='utf-8') as f:
    owasp_content = f.read()

# Read documento_validaciones.tex
with open('documento_validaciones.tex', 'r', encoding='utf-8') as f:
    validaciones_content = f.read()

# Find all MSTG mentions in documents
owasp_controls = set(re.findall(r'MSTG-[A-Z]+-\d+', owasp_content))
validaciones_controls = set(re.findall(r'MSTG-[A-Z]+-\d+', validaciones_content))
all_documented_controls = owasp_controls.union(validaciones_controls)

print('='*80)
print('VERIFICACIÓN DE CONTROLES OWASP MASVS EN DOCUMENTACIÓN')
print('='*80)
print()

print(f'Total controles VERDES (Prioritarios) en checklist: {len(green_controls)}')
print(f'Total controles mencionados en documentos: {len(all_documented_controls)}')
print(f'  - En documento_owasp.tex: {len(owasp_controls)}')
print(f'  - En documento_validaciones.tex: {len(validaciones_controls)}')
print()

# Check which green controls are documented
documented_green = []
missing_green = []

for ctrl in sorted(green_controls):
    if ctrl in all_documented_controls:
        documented_green.append(ctrl)
    else:
        missing_green.append(ctrl)

print('='*80)
print(f'CONTROLES VERDES IMPLEMENTADOS: {len(documented_green)}/{len(green_controls)} ({len(documented_green)*100//len(green_controls)}%)')
print('='*80)
print()

# Group by category
categories = {
    'ARCH': 'Architecture, Design and Threat Modeling',
    'STORAGE': 'Data Storage and Privacy',
    'CRYPTO': 'Cryptography',
    'AUTH': 'Authentication and Session Management',
    'NETWORK': 'Network Communication',
    'PLATFORM': 'Platform Interaction',
    'CODE': 'Code Quality and Build Settings',
    'RESILIENCE': 'Resilience'
}

for cat_key, cat_name in categories.items():
    cat_green = [c for c in green_controls if f'MSTG-{cat_key}' in c]
    cat_documented = [c for c in documented_green if f'MSTG-{cat_key}' in c]

    if cat_green:
        percentage = len(cat_documented) * 100 // len(cat_green) if cat_green else 0
        print(f'{cat_name}:')
        print(f'  Prioritarios: {len(cat_green)} | Implementados: {len(cat_documented)} ({percentage}%)')
        print(f'  {", ".join(cat_documented) if cat_documented else "Ninguno"}')
        print()

print('='*80)
print(f'CONTROLES VERDES NO DOCUMENTADOS: {len(missing_green)}')
print('='*80)
print()

if missing_green:
    for cat_key, cat_name in categories.items():
        cat_missing = [c for c in missing_green if f'MSTG-{cat_key}' in c]
        if cat_missing:
            print(f'{cat_name}:')
            print(f'  {", ".join(cat_missing)}')
            print()

print('='*80)
print('RECOMENDACIONES')
print('='*80)
print()
if len(documented_green) >= len(green_controls) * 0.3:
    print(f'✓ El proyecto cumple con el objetivo mínimo del 30% ({len(documented_green)*100//len(green_controls)}%)')
else:
    print(f'✗ El proyecto NO cumple con el objetivo mínimo del 30% (solo {len(documented_green)*100//len(green_controls)}%)')

print()
print('Controles prioritarios por implementar:')
priority_missing = [c for c in missing_green if any(x in c for x in ['STORAGE', 'CRYPTO', 'AUTH', 'NETWORK'])]
print(f'  - Críticos: {len([c for c in priority_missing if "NETWORK" in c or "CRYPTO" in c])}')
print(f'  - Importantes: {len([c for c in priority_missing if "AUTH" in c or "STORAGE" in c])}')
