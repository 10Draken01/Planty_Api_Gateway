from openpyxl import load_workbook

# Load workbook
wb = load_workbook('Mobile_App_Security_Checklist_es.xlsx')
ws = wb['Security Requirements']

print('=== CONTROLES OWASP MASVS MARCADOS EN VERDE (PRIORIDAD) ===\n')

# Find all MSTG controls with green marker
green_controls = []
current_category = ""

for i, row in enumerate(ws.iter_rows(min_row=1, max_row=200), 1):
    # Check for category in column B
    if row[1].value and 'Requirements' in str(row[1].value):
        current_category = row[1].value
        continue

    # Get MSTG ID from column C (index 2)
    mstg_id = row[2].value if len(row) > 2 else None
    desc = row[3].value if len(row) > 3 else None  # Description in column D

    if mstg_id and str(mstg_id).startswith('MSTG-'):
        # Check column 5 (Col5, index 5) for green color 0099CC00
        is_green = False
        if len(row) > 5 and row[5].fill and hasattr(row[5].fill, 'start_color') and row[5].fill.start_color:
            try:
                color = str(row[5].fill.start_color.rgb) if hasattr(row[5].fill.start_color, 'rgb') else str(row[5].fill.start_color.index)
                # The specific green color in this checklist
                if '0099CC00' in color or '99CC00' in color:
                    is_green = True
            except:
                pass

        if is_green:
            control_info = {
                'id': mstg_id,
                'desc': desc,
                'category': current_category,
                'row': i
            }
            green_controls.append(control_info)

# Print results grouped by category
current_cat = ""
for ctrl in green_controls:
    if ctrl['category'] != current_cat:
        current_cat = ctrl['category']
        print(f'\n{"="*80}')
        print(f'{current_cat}')
        print(f'{"="*80}\n')

    print(f'{ctrl["id"]}')
    if ctrl['desc']:
        desc_clean = str(ctrl['desc']).replace('\n', ' ').strip()
        print(f'  {desc_clean}')
    print()

print(f'\n{"="*80}')
print(f'RESUMEN')
print(f'{"="*80}')
print(f'Total controles VERDES (Prioritarios): {len(green_controls)}')
print(f'\nEstos son los controles que deben implementarse como prioridad según el checklist.')
